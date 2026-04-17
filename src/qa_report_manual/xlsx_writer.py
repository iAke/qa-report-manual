"""Render SiteSnapshot → styled Excel (.xlsx) QA report (multi-sheet workbook)."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import SiteSnapshot

# --- brand ---
BRAND_HEX = "0B5FA5"
ZEBRA_HEX = "F2F6FB"
WHITE = "FFFFFF"
MUTED = "555555"

SEV_COLORS = {
    "S1": "B3261E",  # red
    "S2": "A56400",  # amber
    "S3": "805500",  # yellow
    "S4": "16794A",  # green
}

_thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_cell(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BRAND_HEX)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER


def _body_cell(cell, value, *, bold: bool = False, color: str | None = None,
               fill: str | None = None, wrap: bool = True, align: str = "left") -> None:
    cell.value = value
    cell.font = Font(name="Calibri", size=10, bold=bold, color=color or "222222")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = BORDER


def _write_table(ws, headers: list[str], rows: list[list], *,
                 table_name: str, start_row: int = 1, start_col: int = 1,
                 col_widths: list[int] | None = None,
                 severity_col: int | None = None) -> None:
    # header
    for j, h in enumerate(headers):
        _header_cell(ws.cell(row=start_row, column=start_col + j), h)
    # body
    for i, row in enumerate(rows, start=1):
        zebra = ZEBRA_HEX if i % 2 == 0 else None
        for j, val in enumerate(row):
            fill = zebra
            color = None
            bold = False
            if severity_col is not None and j == severity_col and isinstance(val, str) and val in SEV_COLORS:
                color = SEV_COLORS[val]
                bold = True
            _body_cell(ws.cell(row=start_row + i, column=start_col + j), val,
                       fill=fill, color=color, bold=bold)
    # freeze header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
    # column widths
    if col_widths:
        for idx, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(start_col + idx)].width = w
    # Excel Table for filter/sort
    if rows:
        last_col = get_column_letter(start_col + len(headers) - 1)
        ref = f"{get_column_letter(start_col)}{start_row}:{last_col}{start_row + len(rows)}"
        try:
            tbl = Table(displayName=table_name, ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight1", showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(tbl)
        except Exception:
            pass


def _title_row(ws, title: str, subtitle: str = "") -> None:
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1)
    c.value = title
    c.font = Font(name="Calibri", size=18, bold=True, color=BRAND_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells("A2:H2")
        c2 = ws.cell(row=2, column=1)
        c2.value = subtitle
        c2.font = Font(name="Calibri", size=10, italic=True, color=MUTED)


def _verdict(site: SiteSnapshot) -> tuple[str, str]:
    sev = [f.severity for f in site.findings]
    if any(s == "S1" for s in sev):
        return "FAIL — critical issues block release", SEV_COLORS["S1"]
    if sum(1 for s in sev if s == "S2") >= 2:
        return "CONDITIONAL — major issues need remediation", SEV_COLORS["S2"]
    if sev:
        return "PASS WITH NOTES — minor/moderate issues", SEV_COLORS["S4"]
    return "PASS — no automated red flags (manual QA still required)", SEV_COLORS["S4"]


# ---------------------------------------------------------------------------

def write_qa_report_xlsx(site: SiteSnapshot, out_path: Path) -> None:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb = Workbook()

    pages = site.pages
    ok = [p for p in pages if not p.load_error]
    failed = [p for p in pages if p.load_error]
    total_imgs = sum(p.total_images for p in ok)
    missing_alt = sum(p.images_without_alt for p in ok)
    total_forms = sum(len(p.forms) for p in ok)
    total_console = sum(len(p.console_errors) for p in ok)

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    _title_row(ws, "QA Report — Site crawl",
               f"Entry: {site.entry_url}  ·  Generated: {now}  ·  Standard: ISO/IEC/IEEE 29119-3 §7")

    verdict, color = _verdict(site)
    ws.cell(row=4, column=1, value="Verdict").font = Font(bold=True, size=11)
    vc = ws.cell(row=4, column=2, value=verdict)
    vc.font = Font(bold=True, size=12, color=color)
    ws.merge_cells("B4:F4")

    sev_counts = {"S1": 0, "S2": 0, "S3": 0, "S4": 0}
    for f in site.findings:
        sev_counts[f.severity] = sev_counts.get(f.severity, 0) + 1

    metrics = [
        ["Entry URL", site.entry_url],
        ["Pages crawled", len(pages)],
        ["Pages failed to load", len(failed)],
        ["Total forms detected", total_forms],
        ["Total images", total_imgs],
        ["Images missing alt", missing_alt],
        ["Console errors (sum)", total_console],
        ["Sitemap URLs seeded", site.sitemap_count],
        ["Findings · S1 (critical)", sev_counts["S1"]],
        ["Findings · S2 (major)", sev_counts["S2"]],
        ["Findings · S3 (moderate)", sev_counts["S3"]],
        ["Findings · S4 (minor)", sev_counts["S4"]],
    ]
    _write_table(
        ws, ["Metric", "Value"], metrics,
        table_name="SummaryMetrics", start_row=6, col_widths=[34, 60],
    )

    # Severity legend block
    legend_row = 6 + len(metrics) + 3
    ws.cell(row=legend_row, column=1, value="Severity legend").font = Font(bold=True, size=11, color=BRAND_HEX)
    legend = [
        ["S1", "Production down / data loss / blocker"],
        ["S2", "Major feature broken"],
        ["S3", "Moderate, workaround exists"],
        ["S4", "Minor / cosmetic"],
    ]
    _write_table(ws, ["Severity", "Meaning"], legend,
                 table_name="SeverityLegend", start_row=legend_row + 1,
                 col_widths=[12, 60], severity_col=0)

    # ---------------- Findings ----------------
    ws = wb.create_sheet("Findings")
    _title_row(ws, "Auto-derived findings",
               "Severity-rated issues detected automatically. Each row links to a WCAG / OWASP / perf reference where relevant.")
    rows = [
        [f.severity, f.area, f.message, f.reference or "-", len(f.pages), "; ".join(f.pages[:5])]
        for f in sorted(site.findings, key=lambda x: x.severity)
    ]
    _write_table(
        ws, ["Severity", "Area", "Finding", "Reference", "# Pages", "Sample pages"],
        rows, table_name="Findings", start_row=4,
        col_widths=[10, 16, 60, 28, 8, 60], severity_col=0,
    )

    # ---------------- Per-page ----------------
    ws = wb.create_sheet("Per-page")
    _title_row(ws, "Per-page overview")
    rows = []
    for p in pages:
        if p.load_error:
            rows.append([p.final_url or p.url, "ERROR", "", "", "", "", "", "", "", p.load_error[:300]])
            continue
        h1c = sum(1 for lvl, _ in p.headings if lvl == 1)
        a = p.a11y
        rows.append([
            p.final_url or p.url,
            "OK",
            p.title or "",
            p.lang or "",
            h1c,
            f"{p.images_without_alt}/{p.total_images}",
            len(p.forms),
            len(p.console_errors),
            "/".join([
                "header" if a.has_header else "-",
                "nav" if a.has_nav else "-",
                "main" if a.has_main else "-",
                "footer" if a.has_footer else "-",
            ]),
            "yes" if a.has_viewport_meta else "no",
        ])
    _write_table(
        ws, ["URL", "Status", "Title", "Lang", "H1", "Missing alt / total", "Forms", "Console", "Landmarks", "Viewport meta"],
        rows, table_name="PerPage", start_row=4,
        col_widths=[55, 9, 42, 8, 6, 18, 8, 9, 26, 13],
    )

    # ---------------- Performance ----------------
    ws = wb.create_sheet("Performance")
    _title_row(ws, "Performance snapshot",
               "Core Web Vitals-adjacent metrics captured via Performance API. Budgets: load ≤ 4000 ms, transfer ≤ 3 MB.")
    rows = []
    for p in ok:
        bytes_kb = round(p.perf.transfer_bytes / 1024)
        rows.append([
            p.final_url or p.url,
            p.perf.load_ms,
            p.perf.dom_content_loaded_ms,
            bytes_kb,
            p.perf.resource_count,
            "⚠︎ over budget" if (p.perf.load_ms > 4000 or p.perf.transfer_bytes > 3_000_000) else "ok",
        ])
    _write_table(
        ws, ["URL", "Load (ms)", "DCL (ms)", "Transfer (KB)", "Resources", "Budget"],
        rows, table_name="Perf", start_row=4,
        col_widths=[55, 12, 12, 14, 12, 18],
    )

    # ---------------- Forms ----------------
    ws = wb.create_sheet("Forms")
    _title_row(ws, "Forms detected across site")
    rows = []
    for p in ok:
        for idx, f in enumerate(p.forms, start=1):
            field_str = ", ".join(
                f"{x.label}{'*' if x.required else ''} [{x.type}]" for x in f.fields
            ) or "(no fields)"
            rows.append([p.final_url or p.url, idx, f.method, f.action or "/", len(f.fields), field_str])
    _write_table(
        ws, ["Page", "#", "Method", "Action", "Fields", "Field list (*=required)"],
        rows, table_name="Forms", start_row=4,
        col_widths=[50, 5, 10, 30, 9, 70],
    )

    # ---------------- Console errors ----------------
    ws = wb.create_sheet("Console")
    _title_row(ws, "Console errors (per page, sampled)")
    rows = []
    for p in ok:
        for err in p.console_errors[:20]:
            rows.append([p.final_url or p.url, err[:500]])
    _write_table(
        ws, ["Page", "Message"], rows or [["—", "(none captured)"]],
        table_name="Console", start_row=4, col_widths=[55, 100],
    )

    # ---------------- Checklist ----------------
    ws = wb.create_sheet("Checklist")
    _title_row(ws, "Manual test checklist",
               "Fill Result column with Pass / Fail / Blocked / N/A. Use Notes for evidence links or steps.")
    rows = [
        ["Smoke: every crawled page loads without critical errors", "", "", ""],
        ["Navigation: primary nav consistent across pages", "", "", ""],
        ["Forms: validation, required fields, error messages", "", "", ""],
        ["Links: no broken internal links (sampled)", "", "", ""],
        ["Accessibility: landmarks, heading order, alt text", "", "", ""],
        ["Responsive: mobile / tablet / desktop breakpoints", "", "", ""],
        ["Cross-browser: Chrome / Firefox / Safari", "", "", ""],
        ["Security: HTTPS, cookie flags (if auth)", "", "", ""],
        ["Performance: LCP / CLS spot check", "", "", ""],
        ["i18n / localization (if applicable)", "", "", ""],
    ]
    _write_table(
        ws, ["Area", "Test ID", "Result", "Notes"], rows,
        table_name="Checklist", start_row=4, col_widths=[55, 14, 18, 60],
    )

    # ---------------- Defect log ----------------
    ws = wb.create_sheet("Defects")
    _title_row(ws, "Defect log (template)", "One row per defect. IDs should be stable (e.g. DEF-001).")
    rows = [["DEF-001", "S1–S4", "Priority", "", "", "", "", "", "", "screenshot / HAR link"]]
    _write_table(
        ws,
        ["ID", "Severity", "Priority", "Page URL", "Area", "Summary",
         "Steps to reproduce", "Expected", "Actual", "Evidence"],
        rows, table_name="Defects", start_row=4,
        col_widths=[10, 10, 10, 40, 18, 40, 40, 30, 30, 26],
    )

    # ---------------- Sign-off ----------------
    ws = wb.create_sheet("Sign-off")
    _title_row(ws, "Sign-off")
    rows = [
        ["QA", "", "", ""],
        ["Product", "", "", ""],
        ["Engineering", "", "", ""],
    ]
    _write_table(
        ws, ["Role", "Name", "Date", "Comments"], rows,
        table_name="SignOff", start_row=4, col_widths=[18, 28, 14, 60],
    )

    wb.save(str(out_path))
